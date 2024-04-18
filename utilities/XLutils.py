import openpyxl  #### library name where you have to use to read write data from excelsheet

def ReadData(file,sheetname, rownum, colnumnum):
    workbook = openpyxl.load_workbook(file)  ## file
    sheet = workbook[sheetname]  ## sheet
    return sheet.cell(row=rownum, colnum=colnumnum).value

def WriteData(file, sheetname,rownum, colnumnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, colnum=colnumnum).value = data  ### enter data
    workbook.save(file) ### save file

def RowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row